import xlsxwriter
import openpyxl

# xb = openpyxl.load_workbook('1-12.xlsx')
xb = openpyxl.load_workbook('13-23.xlsx')

# sheet = xb['！A1']
# sheet = xb['！A2']
# sheet = xb['A3']
# sheet = xb['A4']
# sheet = xb['A5']
# sheet = xb['A6']
# sheet = xb['B7']
# sheet = xb['B8']
# sheet = xb['B9']
# sheet = xb['B10']
# sheet = xb['B11']
# sheet = xb['B12']
# sheet = xb['13组']
# sheet = xb['14组']
# sheet = xb['15组']
# sheet = xb['16组']
# sheet = xb['17组']
# sheet = xb['18组']
# sheet = xb['19组']
# sheet = xb['20组']
# sheet = xb['21组']
# sheet = xb['22组']
sheet = xb['23组']

original_list = []

for i in range(1, 5000):
    original_list.append(str(sheet.cell(row=i, column=1).value).replace('\xa0', ' '))

# print(a)

first_handle_list = []
s = ''
for i in range(0, len(original_list)):
    if original_list[i] == 'None':
        first_handle_list.append(s)
        first_handle_list.append('None')
        s = ''
    elif original_list[i].find('2022') == -1:
        s += original_list[i] + ' '
    else:
        temp = original_list[i].split(' ', 1)
        for j in range(0, len(temp)):
            first_handle_list.append(temp[j])
# print(first_handle_list)

second_handle_list = []
for i in range(0, len(first_handle_list)):
    if first_handle_list[i] != '' and first_handle_list[i].find('撤回了') == -1 and \
            first_handle_list[i].find('修改了群名称') == -1 and first_handle_list[i].find('点击添加好友') == -1:
        second_handle_list.append(first_handle_list[i])
# print(second_handle_list)

final_handle_list = ['']
for i in range(0, len(second_handle_list)):
    if final_handle_list[len(final_handle_list) - 1] != second_handle_list[i]:
        final_handle_list.append(second_handle_list[i])
final_handle_list.remove('')
print(final_handle_list)


# 创建工作簿
# file_name = "！A1.xlsx"
# file_name = "！A2.xlsx"
# file_name = "A3.xlsx"
# file_name = "A4.xlsx"
# file_name = "A5.xlsx"
# file_name = "A6.xlsx"
# file_name = "B7.xlsx"
# file_name = "B8.xlsx"
# file_name = "B9.xlsx"
# file_name = "B10.xlsx"
# file_name = "B11.xlsx"
# file_name = "B12.xlsx"
# file_name = "13组.xlsx"
# file_name = "14组.xlsx"
# file_name = "15组.xlsx"
# file_name = "16组.xlsx"
# file_name = "17组.xlsx"
# file_name = "18组.xlsx"
# file_name = "19组.xlsx"
# file_name = "20组.xlsx"
# file_name = "21组.xlsx"
# file_name = "22组.xlsx"
file_name = "23组.xlsx"
workbook = xlsxwriter.Workbook(file_name)
# 创建工作表
# worksheet = workbook.add_worksheet('！A1')
# worksheet = workbook.add_worksheet('！A2')
# worksheet = workbook.add_worksheet('A3')
# worksheet = workbook.add_worksheet('A4')
# worksheet = workbook.add_worksheet('A5')
# worksheet = workbook.add_worksheet('A6')
# worksheet = workbook.add_worksheet('B7')
# worksheet = workbook.add_worksheet('B8')
# worksheet = workbook.add_worksheet('B9')
# worksheet = workbook.add_worksheet('B10')
# worksheet = workbook.add_worksheet('B11')
# worksheet = workbook.add_worksheet('B12')
# worksheet = workbook.add_worksheet('13组')
# worksheet = workbook.add_worksheet('14组')
# worksheet = workbook.add_worksheet('15组')
# worksheet = workbook.add_worksheet('16组')
# worksheet = workbook.add_worksheet('17组')
# worksheet = workbook.add_worksheet('18组')
# worksheet = workbook.add_worksheet('19组')
# worksheet = workbook.add_worksheet('20组')
# worksheet = workbook.add_worksheet('21组')
# worksheet = workbook.add_worksheet('22组')
worksheet = workbook.add_worksheet('23组')

line = []
row = 0
col = 0
i = 0
while i < len(final_handle_list):
    if final_handle_list[i] != 'None':
        line.append(final_handle_list[i])
        # print(line)
        i += 1
    else:
        worksheet.write_row(row, col, line)
        row += 1
        i += 1
        line.clear()

# 关闭工作簿
workbook.close()
